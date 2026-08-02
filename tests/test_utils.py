import io

import openpyxl
import pandas as pd
import pytest

from backend.utils import dataframe_to_excel_bytes


def test_color_scale_columns_add_conditional_formatting():
    df = pd.DataFrame({"A": [10, 50, 90], "B": ["x", "y", "z"]})
    xlsx_bytes = dataframe_to_excel_bytes(df, sheet_name="Sheet1", color_scale_columns=("A",))
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Sheet1"]
    ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    assert any("A2" in r for r in ranges)


def test_color_scale_columns_ignore_unknown_column_names():
    df = pd.DataFrame({"A": [1, 2, 3]})
    # Must not raise even though "DOES_NOT_EXIST" isn't a real column.
    xlsx_bytes = dataframe_to_excel_bytes(df, color_scale_columns=("DOES_NOT_EXIST",))
    assert xlsx_bytes


def test_categorical_color_columns_apply_solid_fill():
    df = pd.DataFrame({"Status": ["Available", "Unavailable"]})
    xlsx_bytes = dataframe_to_excel_bytes(
        df,
        categorical_color_columns={"Status": {"Available": "63BE7B", "Unavailable": "F8696B"}},
    )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws["A2"].fill.start_color.rgb.endswith("63BE7B")
    assert ws["A3"].fill.start_color.rgb.endswith("F8696B")


def test_categorical_color_columns_leaves_unmapped_values_unstyled():
    df = pd.DataFrame({"Status": ["Unknown"]})
    xlsx_bytes = dataframe_to_excel_bytes(
        df, categorical_color_columns={"Status": {"Available": "63BE7B"}}
    )
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws["A2"].fill.fill_type is None


def test_empty_dataframe_with_color_scale_does_not_raise():
    df = pd.DataFrame(columns=["A"])
    xlsx_bytes = dataframe_to_excel_bytes(df, color_scale_columns=("A",))
    assert xlsx_bytes
