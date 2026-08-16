"""
utils.py
--------
Shared, generic helpers with no ATP-specific business logic:
  - logging setup
  - text normalization (used for robust seller-name matching)
  - a lightweight in-memory TTL cache for calculation results
  - an Excel export helper (styled header, % number format, autofit)
  - a small timing context manager
"""
from __future__ import annotations

import io
import logging
import re
import threading
import time
import unicodedata
import uuid
from contextlib import contextmanager
from dataclasses import dataclass, field
from typing import Any, Iterator

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import get_settings

# --------------------------------------------------------------------------- #
# Filename sanitization
# --------------------------------------------------------------------------- #
# Characters genuinely unsafe in a filename across Windows/Linux/macOS: path
# separators, Windows-reserved punctuation, and control characters.
# Everything else — including non-ASCII scripts like Persian/Arabic — is left
# untouched, so seller names stay human-readable in exported filenames (an
# earlier ASCII-only allowlist collapsed every non-Latin seller name, e.g.
# "آذر نقره", down to a bare "unknown").
_FILENAME_UNSAFE_RE = re.compile(r'[\\/:*?"<>|\x00-\x1f]+')


def safe_filename_part(value: Any) -> str:
    text = _FILENAME_UNSAFE_RE.sub("_", str(value)).strip("_. ")
    return text or "unknown"

# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #
def get_logger(name: str) -> logging.Logger:
    """Return a module-level logger configured once from Settings.log_level."""
    logger = logging.getLogger(name)
    if not logger.handlers:
        settings = get_settings()
        handler = logging.StreamHandler()
        handler.setFormatter(
            logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
        )
        logger.addHandler(handler)
        logger.setLevel(settings.log_level)
        logger.propagate = False
    return logger


# --------------------------------------------------------------------------- #
# Text normalization
# --------------------------------------------------------------------------- #
_WHITESPACE_RE = re.compile(r"\s+")

# Arabic-style characters that commonly appear in Persian Excel exports but
# are visually/semantically equivalent to their Persian counterparts.
_ARABIC_TO_PERSIAN = str.maketrans({"ي": "ی", "ك": "ک"})


def normalize_text(value: Any) -> str:
    """
    Normalize a raw cell value into a clean display string.
    Strips surrounding whitespace, collapses internal whitespace runs,
    and unifies Arabic/Persian look-alike characters. Does NOT change case,
    since this is used for display as well as matching.
    """
    if value is None or (isinstance(value, float) and value != value):  # NaN != NaN
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.translate(_ARABIC_TO_PERSIAN)
    text = _WHITESPACE_RE.sub(" ", text).strip()
    return text


def normalize_key(value: Any) -> str:
    """
    Normalize a raw cell value into a matching KEY: same as normalize_text
    plus case-folding. Never shown to the user directly.
    """
    return normalize_text(value).casefold()


_INTEGER_LOOKING_RE = re.compile(r"-?\d+\.0+")


def normalize_id(value: Any) -> str:
    """
    Normalize a raw identifier cell (Seller_ID, DKP, DKPC) into a clean
    matching/display string.

    Handles the common Excel gotcha where an ID column gets upcast to
    float64 because *some* other cell in the column is blank/NaN (e.g.
    20911381.0 instead of "20911381"): an integer-valued float has its
    trailing ".0" stripped before str()-ing, and the same stripping is
    applied if the value arrives as text (e.g. "20911381.0") for the same
    reason. Everything else is delegated to normalize_text.
    """
    if value is None or (isinstance(value, float) and value != value):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = normalize_text(value)
    if _INTEGER_LOOKING_RE.fullmatch(text):
        text = text.split(".")[0]
    return text


# --------------------------------------------------------------------------- #
# Timing
# --------------------------------------------------------------------------- #
@contextmanager
def timer() -> Iterator[dict[str, float]]:
    """
    Usage:
        with timer() as t:
            ... work ...
        elapsed = t["seconds"]
    The dict is mutated in place so the caller can read it after the block.
    """
    state: dict[str, float] = {"seconds": 0.0}
    start = time.perf_counter()
    try:
        yield state
    finally:
        state["seconds"] = time.perf_counter() - start


# --------------------------------------------------------------------------- #
# In-memory TTL result cache
# --------------------------------------------------------------------------- #
@dataclass
class _CacheEntry:
    summary_df: pd.DataFrame
    missing_df: pd.DataFrame
    tail_summary_df: pd.DataFrame
    tail_dkp_list_df: pd.DataFrame
    seller_zip_bytes: bytes | None = None
    tail_seller_zip_bytes: bytes | None = None
    created_at: float = field(default_factory=time.monotonic)


class ResultCache:
    """
    Thread-safe, in-memory, TTL-based cache mapping a result_id to the
    computed DataFrames (Summary, ATP_Missing, Tail_Summary), so the
    download endpoints don't need to recompute the ATP pipeline.

    NOTE: this assumes a single backend process/worker. If the service is
    ever scaled horizontally, replace this with a shared store (Redis, a
    database, or object storage) behind the same put()/get() interface.
    """

    def __init__(self, ttl_seconds: int, max_entries: int) -> None:
        self._ttl = ttl_seconds
        self._max_entries = max_entries
        self._store: dict[str, _CacheEntry] = {}
        self._lock = threading.Lock()

    def put(
        self,
        summary_df: pd.DataFrame,
        missing_df: pd.DataFrame,
        tail_summary_df: pd.DataFrame,
        tail_dkp_list_df: pd.DataFrame,
        seller_zip_bytes: bytes | None = None,
        tail_seller_zip_bytes: bytes | None = None,
    ) -> str:
        result_id = uuid.uuid4().hex
        with self._lock:
            self._evict_expired()
            if len(self._store) >= self._max_entries:
                oldest_id = min(self._store, key=lambda k: self._store[k].created_at)
                del self._store[oldest_id]
            self._store[result_id] = _CacheEntry(
                summary_df=summary_df,
                missing_df=missing_df,
                tail_summary_df=tail_summary_df,
                tail_dkp_list_df=tail_dkp_list_df,
                seller_zip_bytes=seller_zip_bytes,
                tail_seller_zip_bytes=tail_seller_zip_bytes,
            )
        return result_id

    def get(self, result_id: str) -> _CacheEntry | None:
        with self._lock:
            self._evict_expired()
            return self._store.get(result_id)

    def _evict_expired(self) -> None:
        now = time.monotonic()
        expired = [k for k, v in self._store.items() if now - v.created_at > self._ttl]
        for k in expired:
            del self._store[k]


# --------------------------------------------------------------------------- #
# Excel export
# --------------------------------------------------------------------------- #
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Standard Excel "red - yellow - green" 3-color-scale hex codes, matching
# the on-screen background_gradient(cmap="RdYlGn") styling used in the UI.
_COLOR_SCALE_RED = "F8696B"
_COLOR_SCALE_YELLOW = "FFEB84"
_COLOR_SCALE_GREEN = "63BE7B"


def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    percent_columns: tuple[str, ...] = (),
    color_scale_columns: tuple[str, ...] = (),
    categorical_color_columns: dict[str, dict[str, str]] | None = None,
) -> bytes:
    """
    Serialize a DataFrame to a styled .xlsx file in memory.

    Args:
        df: data to export.
        sheet_name: worksheet name.
        percent_columns: column names (as they appear in df) that hold
            0-100 percentage values and should be rendered as "12.34%".
        color_scale_columns: column names that get a red->yellow->green
            conditional-formatting color scale (relative to that column's
            own min/mid/max), mirroring the on-screen background_gradient.
        categorical_color_columns: {column_name: {value_as_str: "RRGGBB"}}
            — cells in that column get a solid fill looked up by their
            exact value (e.g. Tail Badge ST/MT/LT, an Available/
            Unavailable status). Values with no entry are left unstyled.

    Returns:
        Raw xlsx file bytes, ready to stream in an HTTP response.
    """
    categorical_color_columns = categorical_color_columns or {}
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_idx, column_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Percentage number format (values are already 0-100 floats, so we
        # divide by 100 in the display format rather than the data itself).
        for column_name in percent_columns:
            if column_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(column_name) + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = "0.00\\%"

        if len(df) > 0:
            for column_name in color_scale_columns:
                if column_name not in df.columns:
                    continue
                col_idx = df.columns.get_loc(column_name) + 1
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                rule = ColorScaleRule(
                    start_type="min", start_color=_COLOR_SCALE_RED,
                    mid_type="percentile", mid_value=50, mid_color=_COLOR_SCALE_YELLOW,
                    end_type="max", end_color=_COLOR_SCALE_GREEN,
                )
                worksheet.conditional_formatting.add(cell_range, rule)

        for column_name, value_colors in categorical_color_columns.items():
            if column_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(column_name) + 1
            col_letter = get_column_letter(col_idx)
            for row_offset, value in enumerate(df[column_name], start=2):
                hex_color = value_colors.get(str(value))
                if hex_color:
                    worksheet[f"{col_letter}{row_offset}"].fill = PatternFill(
                        start_color=hex_color, end_color=hex_color, fill_type="solid"
                    )

        # Autofit columns (openpyxl has no native autofit; approximate by
        # measuring the longest rendered value per column).
        for col_idx, column_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                [len(str(column_name))]
                + [len(f"{v:.2f}" if isinstance(v, float) else str(v)) for v in df[column_name]]
            )
            worksheet.column_dimensions[col_letter].width = min(max_len + 4, 60)

        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    return buffer.read()
